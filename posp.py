import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.exc import OperationalError
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
import glob
import os
import sys
import unicodedata 
import re
import logging
import cargarpos 
# Logging (salida consola)
# ---------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# ==============================
# 1️⃣ Conexión a PostgreSQL
# ==============================
usuario = 'postgres'
contraseña = 'pasante'
host = 'localhost'
puerto = '5432'
base_datos = 'pospago'

connection_string = f'postgresql://{usuario}:{contraseña}@{host}:{puerto}/{base_datos}'
try:
    engine = create_engine(connection_string)
    with engine.connect() as conn:
        pass
    logging.info("✅ Conexión a PostgreSQL OK.")
except OperationalError as e:
    logging.exception("❌ Error de conexión a PostgreSQL.")
    raise SystemExit(e)



# ==============================
# Función para quitar negrita
# ==============================
def quitar_negrita_excel(ruta_archivo):
    try:
        wb = load_workbook(ruta_archivo)
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold=False)
        wb.save(ruta_archivo)
    except Exception as e:
        logging.exception(f"❌ Error al quitar negrita en {ruta_archivo}: {e}")
        raise

# ==============================
# 2️⃣ Buscar último Excel o CSV
# ==============================
carpeta_base = r'C:\Users\pasante.ti2\Desktop\bases pospago'

archivos_excel = [
    f for f in glob.glob(os.path.join(carpeta_base, "*.xlsx")) +
        glob.glob(os.path.join(carpeta_base, "*.csv"))
    if "_con_anio_mes" not in f and "_incompletos" not in f
]

if not archivos_excel:
    logging.error("❌ No se encontró ningún archivo Excel o CSV original en la carpeta.")
    raise FileNotFoundError("No se encontró ningún archivo Excel o CSV original en la carpeta.")

ruta_base = max(archivos_excel, key=os.path.getmtime)
logging.info(f"📥 Procesando archivo: {ruta_base}")

carpeta_base = r'C:\Users\pasante.ti2\Desktop\bases pospago'
# Supongamos que ya generaste CORRECTA_*.xlsx
ruta_copia = os.path.join(carpeta_base, f"CORRECTA_{datetime.today().month}.xlsx")

# ==============================
# 3️⃣ Leer archivo
# ==============================
try:
    if ruta_base.lower().endswith(".csv"):
        df = pd.read_csv(ruta_base)
    else:
        df = pd.read_excel(ruta_base, sheet_name=0)
except Exception as e:
    logging.exception(f"❌ Error leyendo el archivo {ruta_base}: {e}")
    raise SystemExit(e)

df.columns = [c.lower().strip() for c in df.columns]
logging.info(f"✅ Total de registros cargados: {len(df)}")

# --- Limpieza y normalización de 'desc_forma_pago' 

def limpiar_sin_tildes(texto):
    if pd.isna(texto):
        return ""
    try:
        texto = (
            texto.encode('latin1', errors='ignore')
            .decode('utf-8', errors='ignore')
        )
    except Exception:
        pass

    # Normaliza (quita tildes y diacríticos)
    texto = unicodedata.normalize('NFKD', texto)
    texto = texto.encode('ascii', 'ignore').decode('utf-8')

    # Reemplaza la barra "/" por un espacio
    texto = texto.replace('/', ' ')

    # Elimina caracteres no deseados
    texto = re.sub(r'[^A-Za-z0-9Ññ\s.,-]', '', texto)

    # Elimina espacios múltiples
    texto = re.sub(r'\s+', ' ', texto).strip()

    # Convierte a mayúsculas
    texto = texto.upper()

    return texto


if 'desc_forma_pago' in df.columns:
    df['desc_forma_pago'] = df['desc_forma_pago'].astype(str).apply(limpiar_sin_tildes)
    logging.info("🧹 Columna 'desc_forma_pago' limpiada (sin tildes ni caracteres especiales).")





# ==============================
# 🔟 Año y mes actuales
# ==============================
fecha_actual = datetime.today()
meses = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
    7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}
mes_actual = meses[fecha_actual.month]

# ==============================
# 4️⃣ Validar datos
# ==============================
# blindaciones iniciales: columnas esperadas
for col_exp in ['nombre_completo', 'identificacion', 'celular']:
    if col_exp not in df.columns:
        logging.warning(f"⚠️ Columna esperada '{col_exp}' no encontrada. Se creará vacía.")
        df[col_exp] = ""

df['nombre_completo'] = df.get('nombre_completo', '').fillna('').astype(str)
df['identificacion'] = df.get('identificacion', '').fillna('').astype(str)
df['celular'] = df.get('celular', '').fillna('').astype(str)

mask_incompleto_id_vacia = (df['identificacion'].str.strip() == '') & (df['nombre_completo'].str.strip() != '')
mask_celular_invalido = df['celular'].apply(lambda x: len(''.join(filter(str.isdigit, x))) < 8)
mask_incompletos = mask_incompleto_id_vacia | mask_celular_invalido

df['celular_norm'] = df['celular'].apply(lambda x: ''.join(filter(str.isdigit, x)))
duplicados_cel = df[df.duplicated('celular_norm', keep=False) & (df['celular_norm'] != '')].copy()

# ⚠️ Si hay errores, crear Excel y detener
if mask_incompletos.any() or not duplicados_cel.empty: 
    try:
        nombre_archivo = f"INCORRECTA_{mes_actual}.xlsx" 
        ruta_incompletos = os.path.join(carpeta_base, nombre_archivo) 
        with pd.ExcelWriter(ruta_incompletos, engine='openpyxl') as writer:
            if mask_incompletos.any():
                incompletos = df.loc[mask_incompletos].copy()
                incompletos.to_excel(writer, sheet_name='Incompletos', index=False) 
            if not duplicados_cel.empty:
                duplicados_cel.to_excel(writer, sheet_name='Duplicados_Celular', index=False) 
        quitar_negrita_excel(ruta_incompletos)
        logging.error("🚫 Proceso detenido: se encontraron registros incompletos o duplicados.")
        sys.exit("Proceso detenido por registros incorrectos.")
    except Exception as e:
        logging.exception("❌ Error al generar archivo de registros incorrectos.")
        raise SystemExit(e)

df.drop(columns=['celular_norm'], inplace=True)

# ==============================
# 5️⃣ Normalizaciones y reglas
# ==============================
if 'categoria1' not in df.columns: 
    df['categoria1'] = 'NO REGISTRA' 
else:
    df['categoria1'] = df['categoria1'].fillna('').astype(str).str.strip()
    df.loc[df['categoria1'] == '', 'categoria1'] = 'NO REGISTRA'

for col in df.columns:
    if "categoria" in col:
        df[col] = df[col].fillna("").astype(str).str.strip()
        df.loc[df[col] == "", col] = "NO REGISTRA"

# --- Completar columnas vacías: institucion_financiera, provincia, ciudad ---
for col in ['institucion_financiera', 'provincia', 'ciudad']:
    if col not in df.columns:
        df[col] = 'NO REGISTRA'
    else:
        df[col] = df[col].fillna('').astype(str).str.strip()
        df.loc[df[col] == '', col] = 'NO REGISTRA'
logging.info(" Columnas 'institucion_financiera', 'provincia' y 'ciudad' completadas con 'NO REGISTRA' si estaban vacías.")


mask_nombre_vacio = (df['nombre_completo'].str.strip() == '') & (df['identificacion'].str.strip() != '')
df.loc[mask_nombre_vacio, 'nombre_completo'] = "NO REGISTRA"

for col in df.columns:
    if "ciclo" in col:
        df[col] = df[col].fillna(0)
if "tb" in df.columns:
    df['tb'] = df['tb'].fillna(0)

# ==============================
# 6️⃣ Campos de fecha
# ==============================
df['texto_extraido'] = fecha_actual.strftime("%d%b%Y").lower()
df['año'] = fecha_actual.year
df['mes'] = mes_actual
cols = ['año', 'mes', 'texto_extraido'] + [c for c in df.columns if c not in ['año','mes','texto_extraido']]
df = df[cols]

# ==============================
# 7️⃣ Normalizar celulares
# ==============================
def normalizar_celular(c):
    if pd.isna(c):
        return ""
    c = str(c).strip().replace(".0", "")
    c = "".join(filter(str.isdigit, c))
    if len(c) == 9:
        return "0" + c
    elif len(c) == 8:
        return "09" + c
    return c

if "celular" in df.columns:
    df['celular'] = df['celular'].apply(normalizar_celular)

# ==============================
# 8️⃣ Catálogo de planes
# ==============================
catalogo_path = os.path.join(carpeta_base, "nuevo", "catalogos bases.xlsx")
if not os.path.exists(catalogo_path):
    logging.warning("⚠️ Catálogo de planes no encontrado en 'nuevo/catalogos bases.xlsx'. Se omitirá relleno de descripción.")
    catalogo_df = pd.DataFrame()
else:
    try: 
        catalogo_df = pd.read_excel(catalogo_path)
        catalogo_df.columns = [c.lower().strip() for c in catalogo_df.columns] 
    except Exception as e:
        logging.exception("❌ Error leyendo catálogo de planes.")
        catalogo_df = pd.DataFrame()

desc_col = None
if not catalogo_df.empty:
    desc_col = next((c for c in catalogo_df.columns if "descripcion" in c or "descripción" in c), None)
    if desc_col is None:
        logging.warning("⚠️ No se encontró columna de descripción en el catálogo.")
        desc_col = None

catalogo_dict = {}
if desc_col is not None and 'id_plan' in catalogo_df.columns:
    catalogo_dict = dict(zip(catalogo_df['id_plan'], catalogo_df[desc_col]))

def rellenar_descripcion(row):
    id_plan = row.get('id_plan')
    if pd.notna(id_plan) and id_plan in catalogo_dict:
        return catalogo_dict[id_plan]
    return row.get('descripcion_plan', "") if 'descripcion_plan' in row.index else ""

if 'id_plan' in df.columns and catalogo_dict:
    try:
        df['descripcion_plan'] = df.apply(rellenar_descripcion, axis=1)
    except Exception as e:
        logging.exception("❌ Error rellenando descripciones de plan. Se continuará sin esa información.")

# ==============================
# 9️⃣ Guardar base correcta
# ==============================
try:
    nombre_archivo = f"CORRECTA_{mes_actual}.xlsx"
    ruta_copia = os.path.join(carpeta_base, nombre_archivo)
    df.to_excel(ruta_copia, index=False)
    quitar_negrita_excel(ruta_copia)
    logging.info(f"📂 Base correcta guardada en: {ruta_copia}")
    logging.info(f"✅ Total registros válidos: {len(df)}")
except Exception as e:
    logging.exception("❌ Error guardando archivo CORRECTA_.")
    raise SystemExit(e)


# # ==============================
# # 🔁 10️⃣ Ejecutar cargarpos.py automáticamente
# # ==============================
if os.path.exists(ruta_copia):
    logging.info("🚀 Ejecutando cargarpos.py con la conexión existente...")
    try:
        cargarpos.cargar_datos(engine, ruta_copia)  # Pasamos engine y ruta del Excel
        logging.info("✅ cargarpos.py ejecutado correctamente.")
    except Exception as e:
        logging.exception(f"❌ Error ejecutando cargarpos.py: {e}")
else:
    logging.warning("⚠️ No se encontró el archivo CORRECTA. No se ejecuta cargarpos.py.") 