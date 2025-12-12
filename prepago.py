import pandas as pd # manipulación de datos 
from sqlalchemy import create_engine # manejo de base de datos
from sqlalchemy.exc import OperationalError # Requiere: pip install sqlalchemy 
from datetime import datetime # manejo de fechas
from openpyxl import load_workbook # manipulación de Excel
from openpyxl.styles import Font # estilos de Excel
import glob # manejo de archivos
import os # manejo de sistema operativo
import sys # manejo de sistema
import logging # manejo de logs
import cargacompletapre  # Script de carga
from sqlalchemy.engine.url import URL  # Requiere: pip install sqlalchemy
import tkinter as tk # interfaz gráfica
from tkinter import filedialog # diálogo de archivos
from sqlalchemy import text

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# ========= Conexión a la base de datos (PostgreSQL) =========
usuario = "analista"
contraseña = "2025Anal1st@"   
host = "192.168.10.116"
puerto = 5432
base_datos = "BcorpPrePrueba"

# Requiere: pip install psycopg2-binary
url = URL.create(
    drivername="postgresql+psycopg2",
    username=usuario,
    password=contraseña,
    host=host,
    port=puerto,
    database=base_datos,
)

try:
    engine = create_engine(
        url,
        pool_pre_ping=True,
        pool_size=5,
        max_overflow=10,
        pool_timeout=60,
    )
    with engine.connect() as conn:
        logging.info("✅ Conexión a PostgreSQL OK.")
except OperationalError as e:
    logging.exception("❌ Error de conexión a PostgreSQL.")
    raise SystemExit(e)

# ==============================
# Función para quitar negrita en Excel
# ==============================
def quitar_negrita_excel(ruta_archivo):
    try:
        wb = load_workbook(ruta_archivo)
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold=False)
        wb.save(ruta_archivo)
    except Exception as e:
        logging.exception(f"❌ Error al quitar negrita en {ruta_archivo}: {e}")
        raise

# ==============================
# 2️⃣ Seleccionar archivo manualmente (explorador de archivos)
# ==============================

root = tk.Tk()
root.withdraw()  # Oculta la ventana principal de Tkinter

ruta_base = filedialog.askopenfilename(
    title="Selecciona el archivo Excel o CSV a procesar",
    filetypes=[("Archivos Excel o CSV", "*.xlsx *.csv")]
)

if not ruta_base:
    logging.error("❌ No se seleccionó ningún archivo. Proceso cancelado.")
    raise SystemExit("No se seleccionó ningún archivo.")

logging.info(f"📥 Procesando archivo seleccionado: {ruta_base}")

# ✅ Guardar con el mismo nombre del archivo original, pero con prefijo copia-
carpeta_base = os.path.dirname(ruta_base)

escritorio = os.path.join(os.path.expanduser("~"), "Desktop")

# Carpeta general "copias"
carpeta_general = os.path.join(escritorio, "copias")
os.makedirs(carpeta_general, exist_ok=True)

# Carpeta específica del tipo
carpeta_tipo = os.path.join(carpeta_general, "prepago")
os.makedirs(carpeta_tipo, exist_ok=True)

# Nombre base del archivo original
nombre_original = os.path.splitext(os.path.basename(ruta_base))[0]

# Nombre final de la copia
nombre_copia = f"copia-{nombre_original}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Ruta final donde se guardará la copia
ruta_copia = os.path.join(carpeta_tipo, nombre_copia)
# ==============================
# 3️⃣ Leer archivo
# ==============================
try:
    if ruta_base.lower().endswith(".csv"):
        df = pd.read_csv(ruta_base)
    else:
        df = pd.read_excel(ruta_base, sheet_name=0)
except Exception as e:
    logging.exception(f"❌ Error leyendo el archivo {ruta_base}: {e}")
    raise SystemExit(e)

df.columns = [c.lower().strip() for c in df.columns]
logging.info(f"✅ Total de registros cargados: {len(df)}")

# ==============================
# 4️⃣ Validaciones básicas
# ==============================
for col_exp in ['nombre_completo', 'identificacion', 'celular', 'monto_recarga']:
    if col_exp not in df.columns:
        logging.warning(f"⚠️ Columna esperada '{col_exp}' no encontrada. Se creará vacía.")
        df[col_exp] = "" if col_exp != 'monto_recarga' else 0

df['nombre_completo'] = df.get('nombre_completo', '').fillna('').astype(str)
df['identificacion'] = df.get('identificacion', '').fillna('').astype(str)
df['celular'] = df.get('celular', '').fillna('').astype(str)
df['monto_recarga'] = pd.to_numeric(df.get('monto_recarga', 0), errors='coerce').fillna(0)

mask_incompleto_id_vacia = (df['identificacion'].str.strip() == '') & (df['nombre_completo'].str.strip() != '')
mask_celular_invalido = df['celular'].apply(lambda x: len(''.join(filter(str.isdigit, x))) < 8)
mask_incompletos = mask_incompleto_id_vacia | mask_celular_invalido

df['celular_norm'] = df['celular'].apply(lambda x: ''.join(filter(str.isdigit, x)))
duplicados_cel = df[df.duplicated('celular_norm', keep=False) & (df['celular_norm'] != '')].copy()

if mask_incompletos.any() or not duplicados_cel.empty:
    nombre_archivo = f"INCORRECTA_{datetime.today().month}.xlsx"
    ruta_incompletos = os.path.join(carpeta_base, nombre_archivo)
    with pd.ExcelWriter(ruta_incompletos, engine='openpyxl') as writer:
        if mask_incompletos.any():
            df.loc[mask_incompletos].to_excel(writer, sheet_name='Incompletos', index=False)
        if not duplicados_cel.empty:
            duplicados_cel.to_excel(writer, sheet_name='Duplicados_Celular', index=False)
    quitar_negrita_excel(ruta_incompletos)
    logging.error("🚫 Proceso detenido: registros incorrectos.")
    sys.exit("Proceso detenido por registros incorrectos.")

df.drop(columns=['celular_norm'], inplace=True)

# ==============================
# 5️⃣ Añadir año, mes y texto_extraido en español
# ==============================
fecha_actual = datetime.today()
meses = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
    7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}
mes_actual = meses[fecha_actual.month]

# Crear columnas al inicio
df.insert(0, 'año', fecha_actual.year)
df.insert(1, 'mes', mes_actual)
df.insert(2, 'texto_extraido', fecha_actual.strftime("%d%b%Y").lower())

# ==============================
# 6️⃣ Normalizar celulares
# ==============================
def normalizar_celular(c):
    if pd.isna(c):
        return ""
    c = str(c).strip().replace(".0", "")
    c = "".join(filter(str.isdigit, c))
    if len(c) == 9:
        return "0" + c
    elif len(c) == 8:
        return "09" + c
    return c

df['celular'] = df['celular'].apply(normalizar_celular)

# 7️⃣ Guardar COPIA SOLO en la carpeta generada (ruta_copia)
df.to_excel(ruta_copia, index=False)
quitar_negrita_excel(ruta_copia)
logging.info(f"📂 Base copiada guardada en: {ruta_copia}")

ruta_correcta = ruta_copia 
nombre_archivo = nombre_copia

# ==============================
# 8️⃣ Ejecutar cargacompletapre.py y actualizar nombre_base
# ==============================

if os.path.exists(ruta_correcta):
    logging.info("🚀 Ejecutando cargacompletapre.py con la conexión existente...")

    # Obtener el nombre original SIN el prefijo "copia-" ni extensión
    nombre_sin_prefijo = os.path.splitext(nombre_original.replace("copia-", ""))[0]

    try:
        cargacompletapre.run_cargarpre(engine, ruta_correcta)
        logging.info("✅ cargacompletapre.py ejecutado correctamente.")

        # 🔄 Actualizar el campo nombre_base en la tabla periodo_carga
        with engine.connect() as conn:
            conn.execute(
                text("""
                    UPDATE periodo_carga
                    SET nombre_base = :nombre
                    WHERE id_periodo = (SELECT MAX(id_periodo) FROM periodo_carga)
                """),
                {"nombre": nombre_sin_prefijo}
            )
            conn.commit()
        logging.info(f"🆗 nombre_base actualizado con '{nombre_sin_prefijo}' en periodo_carga.")

    except Exception as e:
        logging.exception(f"❌ Error ejecutando cargarpre.py o actualizando nombre_base: {e}")

else:
    logging.warning("⚠️ No se encontró el archivo copia. No se ejecuta cargacompletapre.py.")
