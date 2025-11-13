import pandas as pd
from sqlalchemy import create_engine, text
from sqlalchemy.exc import OperationalError
from datetime import datetime
from openpyxl import load_workbook 
from openpyxl.styles import Font
import glob
import os
import sys
import unicodedata 
import re
import logging
import cargacompletapyme 

# ==========================================
# 🔧 Configuración de Logging
# ==========================================
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# ==========================================
# 1️⃣ Conexión a PostgreSQL
# ==========================================
usuario = 'postgres'
contraseña = 'pasante'
host = 'localhost'
puerto = '5432'
base_datos = 'base_pyme'

connection_string = f'postgresql://{usuario}:{contraseña}@{host}:{puerto}/{base_datos}'
try:
    engine = create_engine(connection_string)
    with engine.connect() as conn:
        pass
    logging.info("✅ Conexión a PostgreSQL OK.")
except OperationalError as e:
    logging.exception("❌ Error de conexión a PostgreSQL.")
    raise SystemExit(e)

# ==========================================
# 2️⃣ Función para quitar negrita
# ==========================================
def quitar_negrita_excel(ruta_archivo):
    try:
        wb = load_workbook(ruta_archivo)
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold=False)
        wb.save(ruta_archivo)
    except Exception as e:
        logging.exception(f"❌ Error al quitar negrita en {ruta_archivo}: {e}")
        raise

# ==============================
# 2️⃣ Seleccionar archivo manualmente (explorador de archivos)
# ==============================
import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.withdraw()  # Oculta la ventana principal de Tkinter

ruta_base = filedialog.askopenfilename(
    title="Selecciona el archivo Excel o CSV a procesar",
    filetypes=[("Archivos Excel o CSV", "*.xlsx *.csv")]
)

if not ruta_base:
    logging.error("❌ No se seleccionó ningún archivo. Proceso cancelado.")
    raise SystemExit("No se seleccionó ningún archivo.")

logging.info(f"📥 Procesando archivo seleccionado: {ruta_base}")

# ✅ Guardar con el mismo nombre del archivo original, pero con prefijo copia-
carpeta_base = os.path.dirname(ruta_base)
nombre_original = os.path.splitext(os.path.basename(ruta_base))[0]
nombre_copia = f"copia-{nombre_original}.xlsx"
ruta_copia = os.path.join(carpeta_base, nombre_copia)

# ==========================================
# 4️⃣ Leer archivo
# ==========================================
try:
    if ruta_base.lower().endswith(".csv"):
        df = pd.read_csv(ruta_base)
    else:
        df = pd.read_excel(ruta_base, sheet_name=0)
except Exception as e:
    logging.exception(f"❌ Error leyendo el archivo {ruta_base}: {e}")
    raise SystemExit(e)

df.columns = [c.lower().strip() for c in df.columns]
logging.info(f"✅ Total de registros cargados: {len(df)}")

# ==========================================
# 5️⃣ Limpieza y normalización
# ==========================================
def limpiar_sin_tildes(texto):
    if pd.isna(texto):
        return ""
    try:
        texto = (
            texto.encode('latin1', errors='ignore')
            .decode('utf-8', errors='ignore')
        )
    except Exception:
        pass
    texto = unicodedata.normalize('NFKD', texto)
    texto = texto.encode('ascii', 'ignore').decode('utf-8')
    texto = texto.replace('/', ' ')
    texto = re.sub(r'[^A-Za-z0-9Ññ\s.,-]', '', texto)
    texto = re.sub(r'\s+', ' ', texto).strip()
    texto = texto.upper()
    return texto

if 'desc_forma_pago' in df.columns:
    df['desc_forma_pago'] = df['desc_forma_pago'].astype(str).apply(limpiar_sin_tildes)
    logging.info("🧹 Columna 'desc_forma_pago' limpiada (sin tildes ni caracteres especiales).")

# ==========================================
# 6️⃣ Año y mes actuales
# ==========================================
fecha_actual = datetime.today()
meses = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
    7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}
mes_actual = meses[fecha_actual.month]

# ==========================================
# 7️⃣ Validar datos
# ==========================================
for col_exp in ['nombre_completo', 'identificacion', 'celular']:
    if col_exp not in df.columns:
        logging.warning(f"⚠️ Columna esperada '{col_exp}' no encontrada. Se creará vacía.")
        df[col_exp] = ""

df['nombre_completo'] = df.get('nombre_completo', '').fillna('').astype(str)
df['identificacion'] = df.get('identificacion', '').fillna('').astype(str)
df['celular'] = df.get('celular', '').fillna('').astype(str)

mask_incompleto_id_vacia = (df['identificacion'].str.strip() == '') & (df['nombre_completo'].str.strip() != '')
mask_celular_invalido = df['celular'].apply(lambda x: len(''.join(filter(str.isdigit, x))) < 8)
mask_incompletos = mask_incompleto_id_vacia | mask_celular_invalido

df['celular_norm'] = df['celular'].apply(lambda x: ''.join(filter(str.isdigit, x)))
duplicados_cel = df[df.duplicated('celular_norm', keep=False) & (df['celular_norm'] != '')].copy()

if mask_incompletos.any() or not duplicados_cel.empty: 
    try:
        nombre_archivo = f"INCORRECTA_{mes_actual}.xlsx" 
        ruta_incompletos = os.path.join(carpeta_base, nombre_archivo) 
        with pd.ExcelWriter(ruta_incompletos, engine='openpyxl') as writer:
            if mask_incompletos.any():
                incompletos = df.loc[mask_incompletos].copy()
                incompletos.to_excel(writer, sheet_name='Incompletos', index=False) 
            if not duplicados_cel.empty:
                duplicados_cel.to_excel(writer, sheet_name='Duplicados_Celular', index=False) 
        quitar_negrita_excel(ruta_incompletos)
        logging.error("🚫 Proceso detenido: se encontraron registros incompletos o duplicados.")
        sys.exit("Proceso detenido por registros incorrectos.")
    except Exception as e:
        logging.exception("❌ Error al generar archivo de registros incorrectos.")
        raise SystemExit(e)

df.drop(columns=['celular_norm'], inplace=True)

# ==========================================
# 8️⃣ Normalizaciones y campos faltantes
# ==========================================
if 'categoria1' not in df.columns: 
    df['categoria1'] = 'NO REGISTRA' 
else:
    df['categoria1'] = df['categoria1'].fillna('').astype(str).str.strip()
    df.loc[df['categoria1'] == '', 'categoria1'] = 'NO REGISTRA'

for col in df.columns:
    if "categoria" in col:
        df[col] = df[col].fillna("").astype(str).str.strip()
        df.loc[df[col] == "", col] = "NO REGISTRA"

for col in ['institucion_financiera', 'provincia', 'ciudad']:
    if col not in df.columns:
        df[col] = 'NO REGISTRA'
    else:
        df[col] = df[col].fillna('').astype(str).str.strip()
        df.loc[df[col] == '', col] = 'NO REGISTRA'
logging.info(" Columnas 'institucion_financiera', 'provincia' y 'ciudad' completadas con 'NO REGISTRA' si estaban vacías.")

mask_nombre_vacio = (df['nombre_completo'].str.strip() == '') & (df['identificacion'].str.strip() != '')
df.loc[mask_nombre_vacio, 'nombre_completo'] = "NO REGISTRA"

for col in df.columns:
    if "ciclo" in col:
        df[col] = df[col].fillna(0)
if "tb" in df.columns:
    df['tb'] = df['tb'].fillna(0)

# ==========================================
# 9️⃣ Campos de fecha
# ==========================================
df['texto_extraido'] = fecha_actual.strftime("%d%b%Y").lower()
df['año'] = fecha_actual.year
df['mes'] = mes_actual
cols = ['año', 'mes', 'texto_extraido'] + [c for c in df.columns if c not in ['año','mes','texto_extraido']]
df = df[cols]

# ==========================================
# 🔟 Normalizar celulares
# ==========================================
def normalizar_celular(c):
    if pd.isna(c):
        return ""
    c = str(c).strip().replace(".0", "")
    c = "".join(filter(str.isdigit, c))
    if len(c) == 9:
        return "0" + c
    elif len(c) == 8:
        return "09" + c
    return c

if "celular" in df.columns:
    df['celular'] = df['celular'].apply(normalizar_celular)

# ==========================================
# 11️⃣ Catálogo de planes
# ==========================================
catalogo_path = r"C:\Users\pasante.ti2\Desktop\bases pospago\nuevo\catalogos bases.xlsx"
if not os.path.exists(catalogo_path):
    logging.warning("⚠️ Catálogo de planes no encontrado en 'nuevo/catalogos bases.xlsx'. Se omitirá relleno de descripción.")
    catalogo_df = pd.DataFrame()
else:
    try: 
        catalogo_df = pd.read_excel(catalogo_path)
        catalogo_df.columns = [c.lower().strip() for c in catalogo_df.columns] 
    except Exception as e:
        logging.exception("❌ Error leyendo catálogo de planes.")
        catalogo_df = pd.DataFrame()

desc_col = None
if not catalogo_df.empty:
    desc_col = next((c for c in catalogo_df.columns if "descripcion" in c or "descripción" in c), None)
    if desc_col is None:
        logging.warning("⚠️ No se encontró columna de descripción en el catálogo.")
        desc_col = None

catalogo_dict = {}
if desc_col is not None and 'id_plan' in catalogo_df.columns:
    catalogo_dict = dict(zip(catalogo_df['id_plan'], catalogo_df[desc_col]))

def rellenar_descripcion(row):
    id_plan = row.get('id_plan')
    if pd.notna(id_plan) and id_plan in catalogo_dict:
        return catalogo_dict[id_plan]
    return row.get('descripcion_plan', "") if 'descripcion_plan' in row.index else ""

if 'id_plan' in df.columns and catalogo_dict:
    try:
        df['descripcion_plan'] = df.apply(rellenar_descripcion, axis=1)
    except Exception as e:
        logging.exception("❌ Error rellenando descripciones de plan. Se continuará sin esa información.")

# ==========================================
# 12️⃣ Guardar base copia-nombreoriginal
# ==========================================
try:
    df.to_excel(ruta_copia, index=False)
    quitar_negrita_excel(ruta_copia)
    logging.info(f"📂 Base guardada como: {ruta_copia}")
    logging.info(f"✅ Total registros válidos: {len(df)}")
except Exception as e:
    logging.exception("❌ Error guardando archivo copia-.")
    raise SystemExit(e)

# ==========================================
# 13️⃣ Ejecutar cargacompletapyme y guardar nombre_base
# ==========================================
if os.path.exists(ruta_copia):
    logging.info("🚀 Ejecutando cargacompletapyme.py con la conexión existente...")
    try:
        cargacompletapyme.cargar_datos(engine, ruta_copia)

        # Guardar el nombre original en periodo_carga.nombre_base
        with engine.connect() as conn:
            conn.execute(
                text("""
                    UPDATE periodo_carga
                    SET nombre_base = :nombre
                    WHERE id_periodo = (SELECT MAX(id_periodo) FROM periodo_carga)
                """),
                {"nombre": nombre_original}
            )
            conn.commit()
        logging.info(f"🗄️ Nombre de la base '{nombre_original}' guardado en periodo_carga.nombre_base.")
        logging.info("✅ cargacompletapyme.py ejecutado correctamente.")
    except Exception as e:
        logging.exception(f"❌ Error ejecutando cargacompletapyme.py o actualizando nombre_base: {e}")
else:
    logging.warning("⚠️ No se encontró el archivo copia-. No se ejecuta cargacompletapyme.py.")
 